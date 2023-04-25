import openpyxl
import pandas as pd
from pyautogui import size
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import subprocess
import shutil
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from bs4 import BeautifulSoup
import time
import datetime
import pyautogui
import pyperclip
import csv
import sys
import os
import math
import requests
import re
import random
import chromedriver_autoinstaller
from PyQt5.QtWidgets import QWidget, QApplication,QTreeView,QFileSystemModel,QVBoxLayout,QPushButton,QInputDialog,QLineEdit,QMainWindow,QMessageBox,QFileDialog,QTextEdit
from PyQt5.QtCore import QCoreApplication
from selenium.webdriver import ActionChains
from datetime import datetime,date,timedelta
import numpy
import datetime
from window import Ui_MainWindow
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
import pybase64
import bcrypt
import http.client
import json
import pprint
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


def get_token(price, productNo, api_id,api_pw):
    time_now = datetime.datetime.now() - datetime.timedelta(seconds=3)
    time_now_stamp = math.ceil(datetime.datetime.timestamp(time_now) * 1000)
    # print(time_now)
    # print(time_now_stamp)

    clientId = api_id  # client id
    clientSecret = api_pw  # client pw
    # clientId=self.clientid
    # clientSecret=self.clientkey
    # timestamp = 1643961623299
    timestamp = time_now_stamp
    # 밑줄로 연결하여 password 생성
    password = clientId + "_" + str(timestamp)
    # bcrypt 해싱
    hashed = bcrypt.hashpw(password.encode('utf-8'), clientSecret.encode('utf-8'))
    # base64 인코딩
    result = pybase64.standard_b64encode(hashed).decode('utf-8')
    # print(result)
    params = {
        "client_id": clientId,
        "timestamp": time_now_stamp,
        "client_secret_sign": result,
        "grant_type": "client_credentials",
        "type": "SELF"
    }
    res = requests.post('https://api.commerce.naver.com/external/v1/oauth2/token', params=params)
    res.raise_for_status()

    token = eval(res.text)['access_token']
    conn = http.client.HTTPSConnection("api.commerce.naver.com")
    headers = {'Authorization': "Bearer {}".format(token)}
    conn.request("GET", "/external/v2/products/channel-products/{}".format(productNo), headers=headers)
    res = conn.getresponse()
    data = res.read()

    result = data.decode("utf-8")


    json_new_result = json.loads(result)
    print("변경할가격은:",price)
    json_new_result['originProduct']['salePrice']=price

    file_path = 'result.json'
    with open(file_path, 'w') as f:
        json.dump(json_new_result, f)

    token_path = 'token.txt'
    f = open(token_path, 'w')
    f.write(token)
    f.close()
    print("겟토큰완료")
def change_price(productNo):
    token_path = 'token.txt'
    with open(token_path) as f:
        lines = f.readlines()
        token = lines[0].strip()

    file_path = 'result.json'
    with open(file_path, 'r') as f:
        data = json.load(f)

    headers = {
        'Authorization': token,
        'content-type': "application/json"
    }

    # pprint.pprint(data)
    print("PUT요청 보내기")
    res = requests.put(
        'https://api.commerce.naver.com/external/v2/products/channel-products/{}'.format(productNo),
        data=json.dumps(data), headers=headers)
    print("PUT요청 완료")
    # res.raise_for_status()
    result = res.status_code
    print('result:', result)
def find_price(productNo):
    token_path = 'token.txt'
    with open(token_path) as f:
        lines = f.readlines()
        token = lines[0].strip()


    file_path = 'result.json'
    with open(file_path, 'r') as f:
        data = json.load(f)
    # pprint.pprint(data)

    headers = {'Authorization': "Bearer {}".format(token)}
    print("111")
    res = requests.get(
        'https://api.commerce.naver.com/external/v2/products/channel-products/{}'.format(productNo),
        headers=headers)
    res.raise_for_status()
    print("111")
    res_dic = json.loads(res.text)
    name = res_dic['originProduct']['name']
    print("이름은:", name)
    try:
        discount_price = int(
            res_dic['originProduct']['customerBenefit']['immediateDiscountPolicy']['mobileDiscountMethod']['value'])
        price = int(json.loads(res.text)['originProduct']['salePrice'] - discount_price)
    except:
        price = int(json.loads(res.text)['originProduct']['salePrice'])

    return name, price
def load_excel(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    no_row = ws.max_row
    print("행갯수:", no_row)
    info_list = []
    for i in range(2, no_row + 1):
        print(i,"번째 행 정보 가져오는 중...")
        productNo = ws.cell(row=i, column=1).value
        if productNo=="" or productNo==None:
            break
        name = ws.cell(row=i, column=2).value
        url_list= ws.cell(row=i, column=3).value

        url_list=url_list.split(",")

        price_low = ws.cell(row=i, column=4).value
        price_tic = int(ws.cell(row=i, column=5).value)
        switch = ws.cell(row=i, column=6).value

        info = [productNo, name,url_list,price_low,price_tic,switch]
        info_list.append(info)

    print("info_list:",info_list)
    return info_list
def get_catalog_price(url, store_name):
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36"}

    while True:
        try:
            res = requests.get(url, headers=headers)
            res.raise_for_status()
            position_fr = res.text.find("{")
            position_rr = res.text.rfind("}")
            result_raw = res.text[position_fr:position_rr + 1]
            result = json.loads(result_raw)
            result_list = result['props']['pageProps']['dehydratedState']['queries']

            mall_total_list = []
            mall_useless=['11번가','G마켓','옥션','쿠팡','위메프','롯데','템스윈공식몰','인터파크','인터파크쇼핑']
            for index, result_elem in enumerate(result_list):
                try:
                    mall_list = result_elem['state']['data']['pages'][0]['products']
                except:
                    # print("없음")
                    mall_list = []
                for mall_elem in mall_list:
                    if mall_elem['mallName'] in mall_useless:
                        continue
                    print("몰이름:", mall_elem['mallName'], "가격:", mall_elem['mobilePrice'])
                    data = [mall_elem['mallName'], int(mall_elem['mobilePrice'])]
                    mall_total_list.append(data)
            print("mall_total_list:", mall_total_list)

            first_flag = True
            for mall_total_elem in mall_total_list:
                price_mall = mall_total_elem[1]
                name_mall = mall_total_elem[0]
                print("몰가격:", price_mall, "몰이름:", name_mall)
                if first_flag == True:
                    least_price = price_mall
                    if name_mall.find(store_name) >= 0:
                        is_first = True
                        print("1등여부:", is_first)
                    else:
                        is_first = False
                    first_flag = False
                elif first_flag == False:
                    second_price = price_mall
                    break
            break
        except:
            print("에러")
            time.sleep(10)
    return least_price, second_price, is_first
def get_target_price(url):
    # url = 'https://smartstore.naver.com/1cc/products/7190863120?NaPm=ct%3Dlfm3pj5k%7Cci%3D743a40b6df75b561265ff23978ea1f990e632c4a%7Ctr%3Dslsc%7Csn%3D4367970%7Chk%3D3b9234ab4ccb9ace4a557ccedc0848348b46b343'
    headers = {
        'User-Agent': 'Mozilla/5.0 (iPhone; CPU iPhone OS 9_3_2 like Mac OS X) AppleWebKit/601.1.46 (KHTML, like Gecko) Version/9.0 Mobile/13F69 Safari/601.1',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Accept-Charset': 'ISO-8859-1,utf-8;q=0.7,*;q=0.3',
        'Accept-Encoding': 'none',
        'Accept-Language': 'en-US,en;q=0.8',
        'Connection': 'keep-alive'}
    res = requests.get(url, headers=headers)
    res.raise_for_status()
    soup = BeautifulSoup(res.text, 'lxml')
    head = soup.find('head')
    script = head.find_all('script')[0]
    position_fr = str(script).find("{")
    position_rr = str(script).rfind("}")
    result_raw = str(script)[position_fr:position_rr + 1]
    result = int(json.loads(result_raw)['offers']['price'])
    print("타겟가격:",result)
    return result

class Thread(QThread):
    # 초기화 메서드 구현
    user_signal = pyqtSignal(str)  # 사용자 정의 시그널 2 생성

    def __init__(self, parent,file_path,store_name,api_id,api_pw):  # parent는 WndowClass에서 전달하는 self이다.(WidnowClass의 인스턴스)
        super().__init__(parent)
        self.parent = parent  # self.parent를 사용하여 WindowClass 위젯을 제어할 수 있다.
        self.running_flag=True
        self.file_path=file_path
        self.store_name=store_name
        self.api_id=api_id
        self.api_pw=api_pw


    def run(self):
        widen_flag=True
        while True:

            # -----------------실행부위-------------
            info_list = load_excel(self.file_path)
            print("엑셀읽어오기완료")
            wb=openpyxl.Workbook()
            ws=wb.active

            column_name=[]
            product_name = []
            column_name.append("시간")
            product_name.append("")
            for info_elem in info_list:
                column_name.append(info_elem[0])
            print('column_name:',column_name)
            for info_elem in info_list:
                product_name.append(info_elem[1])
            print('product_name:',product_name)
            ws.append(column_name)
            ws.append(product_name)


            time_now_first = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            wb.save('price_history_{}.xlsx'.format(time_now_first))

            while True:
                print("조정시작")
                time_now=datetime.datetime.now().strftime("%Y년%m월%d일%H시%M분%S초")
                price_result_list=[]
                price_target_list=[]
                price_result_list.append(time_now)
                price_target_list.append("타겟팅 최저가격")
                for index, info in enumerate(info_list):
                    print("작동여부:",self.running_flag)
                    if self.running_flag==False:
                        break
                    # info = [productNo, name, url_list, price_low, price_tic, switch]
                    productNo = info[0]
                    name = info[1]
                    url_list = info[2]
                    price_low=info[3]
                    price_tic = info[4]
                    switch = info[5]

                    if switch==0:
                        print("조정안하는 상품 스킵함")
                        price_target_list.append("")
                        price_result_list.append("")
                        continue

                    # if len(url_list) == 0:
                    #     print("url없어서 넘어감")
                    #     continue

                    nownow = datetime.datetime.now()
                    nownow = nownow.strftime("%Y%m%d_%H%M")
                    print("★★★★★★★★★★★★★★★★★★★★★★★★★★★")
                    text="{}번째 상품 크롤링 중.. 번호 : {}".format(index+1,productNo)
                    print(text)
                    self.user_signal.emit(text)

                    if switch==1:
                        print("조절가능")
                    elif switch==0:
                        print("조절불가능")
                        continue

                    if productNo == "" or productNo == None:
                        print("상품번호없어서 넘어감")
                        continue



                    while True:
                        try:
                            target_price_list=[]
                            for url_elem in url_list:
                                target_price=get_target_price(url_elem)
                                print('target_price:',target_price)
                                target_price_list.append(target_price)
                                time.sleep(0.2)
                            target_price_list.sort()
                            print('target_price_list_sorted:',target_price_list)
                            price_least=int(target_price_list[0])
                            break
                        except:
                            print("에러로 한번건너뜀1")
                            time.sleep(10)


                    print("토큰발행")
                    get_token(price_least-price_tic, productNo, self.api_id, self.api_pw)
                    print("현재가격찾기")
                    name, current_price = find_price(productNo)
                    print("현재가격찾기완료")

                    if switch==1: #최저가와 상관없이 변경
                        while True:
                            try:
                                price_change=price_least-price_tic
                                if price_least - price_tic < price_low:
                                    price_change=price_low
                                # get_token(price_change,productNo,self.api_id,self.api_pw)
                                break
                            except:
                                print("에러로 한번 건너뜀2")
                                time.sleep(10)


                    elif switch==2: #최저가보다 낮으면 변경 안함
                        if current_price<=price_least:
                            if current_price<price_least-price_tic:
                                print("가격이 너무 낮아서 최저가-틱으로 상향")
                                price_change = price_least - price_tic
                                get_token(price_change, productNo, self.api_id, self.api_pw)
                            else:
                                price_change = price_least - price_tic
                                print("이미 최저가보다 낮아서 변경 안함")
                                price_target_list.append(price_least)
                                price_result_list.append(current_price)
                                continue
                        if current_price>price_least:
                            print("최저가 보단 높으므로 변경")
                            while True:
                                try:

                                    price_change = price_least - price_tic
                                    if price_least - price_tic < price_low:
                                        price_change = price_low
                                    get_token(price_change, productNo, self.api_id, self.api_pw)
                                    break
                                except:
                                    print("에러로 한번 건너뜀2")
                                    time.sleep(10)



                    # name, current_price = find_price(productNo)
                    print("최저가격은:", price_least,"변경할가격:",price_change,"하한가:",price_low)
                    change_price(productNo)
                    time.sleep(0.5)
                    price_target_list.append(price_least)
                    price_result_list.append(price_change)
                    print("-----------------------------------")
                ws.append(price_target_list)
                ws.append(price_result_list)

                if widen_flag==True:
                    for column_cells in ws.columns:
                        length = max(len(str(cell.value)) * 1.1 for cell in column_cells)
                        ws.column_dimensions[column_cells[0].column_letter].width = 17
                        ## 셀 가운데 정렬
                        # for cell in ws[column_cells[0].column_letter]:
                        #     cell.alignment = Alignment(horizontal='center')
                    widen_flag=False
                wb.save('price_history_{}.xlsx'.format(time_now_first))





    def stop(self):
        self.running_flag=False
        self.quit()
        print("1111213123",self.running_flag)


class Example(QMainWindow,Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.path="C:"
        self.index=None
        self.setupUi(self)
        self.setSlot()
        self.show()
        self.file_path=""
        self.running_flag=True
        QApplication.processEvents()





    def start(self):
        self.api_id=self.lineEdit_4.text()
        self.api_pw=self.lineEdit_5.text()
        self.file_path = self.lineEdit_3.text()

        print("11")
        if len(self.file_path)==0:
            QMessageBox.information(self, "에러", "엑셀 파일을 Import 하세요")
        else:
            self.x=Thread(self, self.file_path,self.store_name,self.api_id,self.api_pw)
            self.x.user_signal.connect(self.slot1)  # 사용자 정의 시그널2 슬롯 Connect
            self.x.start()
    def stop(self):
        QCoreApplication.instance().quit()

    def slot1(self, data1):  # 사용자 정의 시그널1에 connect된 function
        self.textEdit.append(str(data1))

    def setSlot(self):
        pass
    def setIndex(self,index):
        pass
    def quit(self):
        QCoreApplication.instance().quit()
    def search(self):
        fname = QFileDialog.getOpenFileName(self, "Open file", './')
        print(fname[0])
        self.file_path=fname[0]
        self.lineEdit_3.setText(fname[0])
        wb=openpyxl.load_workbook(fname[0])
        ws=wb.active
        self.id=ws.cell(row=2,column=8).value
        self.pw=ws.cell(row=2,column=9).value
        self.store_name = ws.cell(row=2, column=11).value
        self.lineEdit_4.setText(self.id)
        self.lineEdit_5.setText(self.pw)
app=QApplication([])
ex=Example()
sys.exit(app.exec_())



if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = MyApp()
    sys.exit(app.exec_())
